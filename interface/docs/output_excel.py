# ! /usr/bin/env python
# coding=utf-8

import os
import sys
import pathlib

try:
    # try to use __file__
    excel_dir = os.path.dirname(os.path.abspath(__file__))
except NameError:
    # fallback to sys.argv[0] or current working directory
    if len(sys.argv) > 0:
        excel_dir = os.path.dirname(os.path.abspath(sys.argv[0]))
    else:
        excel_dir = os.getcwd()

log_dir = pathlib.Path(excel_dir).parent.parent/"log"

if not log_dir.exists():
    log_dir.mkdir(parents=True, exist_ok=True)


from interface.cui_logger import logger as log
from interface.cui_colors import color, color_hex

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Color, Border, Side, Alignment
from openpyxl.styles.fonts import Font
from openpyxl.drawing.image import Image
from tabulate import tabulate as tb

import datetime


def log_wrapping(header, message):
    log.forcedLog(f"[{header} {sys._getframe(2).f_code.co_name}] {message}")


class style:

    font_calibri     = "Calibri"
    font_abadi       = "Abadi"
    font_arial       = "Arial"
    font_1underline  = "single"
    font_2underline  = "double"
    font_subscript   = "subscript"
    font_superscript = "superscript"
    font_basic       = "baseline"

    border_thin  = "thin"
    border_thick = "thick"

    vertical_top     = "top"
    vertical_cener   = "center"
    horizontal_left  = "left"
    horizontal_right = "right"

    header = Font(
        name      = font_calibri,
        size      = 10,
        bold      = True,
        italic    = False,
        vertAlign = font_basic,
        underline = None,
        strike    = False,
        color     = color_hex.white
    )

    context = Font(
        name      = font_calibri,
        size      = 10,
        bold      = False,
        italic    = False,
        vertAlign = font_basic,
        underline = None,
        strike    = False,
        color     = color_hex.black
    )


class excel_frame:

    '''
    00. common
        - self.filename
        - self.wb : workbook
        - self.ws_list
        - self.ws_active : activated worksheet
        - self._row
        - self._col
        - self.set_cell_color : need to be asigned from color_hex class
        - self.active_cell : assigned from coordinate() function
    01. def __init__()
        - filename = None   : generate today.xlsx
        - filename = String : load or generate
        - assign the activated sheet to the property
    02. save : property, manual save
    03. worksheet_add : decorator, add only
    04. worksheet_list : property, return all worksheets name
    05. select_worksheet
        - decorator
        - return activated worksheet title
        - select worksheet
    06. worksheet_title
        - decorator
        - color : refered to self.set_cell_color
        - return activated worksheet title
        - change the title
    07. select_worksheet()
    08. cell : property, return the cell value
    09. cell = value : decorator, add the value
    10. row : decorator, give a number
    11. column : decorator, give a number
    12. image : decorator, require list [object, width, height]
    13. cell_color : decorator, require color_hex.variable
    14. cell_font : property
    15. border : decorator, require color_hex.variable
    16. border_none : set none type border property
    17. vertical_align : decorator, call the style.vertical_xx
    18. horizontal_align : decorator, call style.horizontal_xx
    19. cell_height : decorator, set the pixel number

    
    usage example:

    - xl = excel_frame(file="test") # generate or load the test.xlsx file
    - xl.preset
    - xl.worksheet_list
    - xl.worksheet_add = "temp"
    - xl.select_worksheet = "Sheet"
    - xl.worksheet_title = "test"
    - xl.cell = 2024
    - xl.row = 2
    - xl.column = 4
    - xl.position = 4, 5
    - xl.insert_row
    - xl.insert_column
    - xl.image = ["./log/test.png", 100, 200]
    - xl.set_font_size = 20
    - xl.set_font_color = color_hex.white 
    - xl.cell_font
    - xl.cell_color = color_hex.blue
    - xl.border = color_hex.red
    - xl.border_none
    - xl.vertical_align = style.vertical_top
    - xl.horizontal_align = style.horizontal_right
    - xl.cell_height = 40
    - xl.cell_width = ["B", 1000]

    xl = excel_frame(file="test")
    xl.position = 3,5
    xl.preset
    xl.set_font_bold = True
    xl.set_font_color = color_hex.red
    xl.set_font_type = style.font_abadi
    xl.update_font
    xl.cell = "data"
    xl.close

    xl.position = 6, n
    xl.cell = "header comment"
    xl.set_header
    '''

    def __init__(self, file=None):

        # log.initLogger(log.info)

        if file is None:
            now = datetime.datetime.now()
            self.filename = now.strftime("%Y-%m-%d_%Hh%Mm%Ss") + ".xlsx"
            self.wb = Workbook() # create a new file with single worksheet
        else:
            if file.endswith(".xlsx"):
                self.filename = file
            else:
                self.filename = file + ".xlsx"

            if self.filename in os.listdir(log_dir):
                self.wb = load_workbook(filename=log_dir/self.filename)
            else:
                log_wrapping("excel_log", f"{self.filename} is not in target directory, file has been created")
                self.wb = Workbook()
                self.save

        self.ws_list = self.wb.sheetnames
        self.ws_active = self.wb.active

        self._row = 1
        self._col = 1
        self._tab_color = color_hex.light_green
        self._image  = None
        self._border = style.border_thin
        self._align_vertical   = style.vertical_cener
        self._align_horizontal = style.horizontal_left
        self.cell_active = self.ws_active.cell(row=self._row, column=self._col)

        self.set_cell_color = color_hex.black
        self.set_font_type = style.font_arial
        self.set_font_size = 10
        self.set_font_bold = False
        self.set_font_italic = False
        self.set_font_vert = style.font_basic
        self.set_font_underline = None
        self.set_font_strike = False
        self.set_font_color = color_hex.black

        self.update_font

        log_wrapping("excel_log", f"open {color.red}{self.filename}{color.end} and activate {color.red}{self.ws_active.title}{color.end} worksheet")
        log_wrapping("excel_log", "manual save is required after finish the job")
    
    
    @property
    def close(self):

        self.save
        self.wb.close()
        log_wrapping("excel_log", f"close the workbook {self.filename}")


    @property
    def preset(self):

        header = ["property", "value"]
        ret_map = [header] + [
            ["filename (self.filename)", self.filename],
            ["working sheet (self.ws_active.title)", self.ws_active.title],
            ["assigned row (self._row)", self._row],
            ["assigned column (self._col)", self._col],
            ["tab color (self._tab_color)", self._tab_color],
            ["border style (self._border)", self._border],
            ["cell color (self.set_cell_color)", self.set_cell_color],
            ["vertical align (self._align_vertical)", self._align_vertical],
            ["horizontal align (self._align_horizontal)", self._align_horizontal]
            ]
        print(tb(ret_map, headers="firstrow"))
    

    @property
    def save(self):

        self.wb.save(filename=log_dir/self.filename)


    @property
    def worksheet_list(self):

        # empty the list to update worksheet titles
        self.ws_list.clear()

        '''
        ws_list = self.wb.worksheets
        log.forcedLog("list of worksheets")

        for ws in ws_list:
            match = re.search(r'"([^"]+)"', str(ws))
            if match:
                log.forcedLog(f"{color.blue}- {match.group(1)}{color.end}")
                self.ws_list.append(match.group(1))
        '''

        self.ws_list = self.wb.sheetnames

        log_wrapping("excel_log", "worksheets list")
        for ws in self.ws_list:
            log_wrapping("excel_log", f"{color.blue}- {ws}{color.end}")
    

    @property
    def worksheet_add(self):
        pass


    @worksheet_add.setter
    def worksheet_add(self, title):
        
        # insert at the end by default
        # (option) insert at first position : create_sheet(0)

        self.wb.create_sheet(title=title)
        log_wrapping("excel_log", f"{color.red}{self.wb.worksheets[-1]}{color.end} is added at the end of workbook")
        
        self.select_worksheet = title
        self.save
    

    @property
    def select_worksheet(self):

        log_wrapping("excel_log", f"activated worksheet is {color.red}{self.ws_active.title}{color.end}")
        return self.ws_active.title


    @select_worksheet.setter
    def select_worksheet(self, title):

        previous_ws = self.ws_active.title
        self.ws_active = self.wb[title]
        log_wrapping("excel_log", f"change the working worksheet from {previous_ws} to {color.red}{self.ws_active.title}{color.end}")
    

    @property
    def worksheet_title(self):

        # return the title of current worksheet
        return self.ws_active.title


    @worksheet_title.setter
    def worksheet_title(self, title):

        # remove all restricted characters from variable "title"
        # remove any kinds of leading or trailing apostrophes
        # truncate the title to 31 characters if it exceeds the length limitation

        # change the title from activated worksheet
        # need to preset the self._tab_color
        
        restricted_chars = [':', '\\', '/', '*', '?', '[', ']', "'"]
        for correction in restricted_chars:
            title = title.replace(correction, '')
        title = title[:31]
        title = title.strip("'")

        previous_title = self.ws_active.title
        self.ws_active.title = title
        self.ws_active.sheet_properties.tabColor = self._tab_color
        log_wrapping("excel_log", f"update worksheet title from {color.blue}{previous_title}{color.end} to {color.red}{self.ws_active.title}{color.end}")
        self.save


    @property
    def cell(self):

        # require setting row and column
        # return the value
        return self.cell_active.value


    @cell.setter
    def cell(self, value):

        # require setting row and column
        self.cell_active.value = value
        self.save
    

    @property
    def row(self):
        return self._row
    

    @row.setter
    def row(self, n):

        # if len(list(filter(lambda x: x <= 0, coorinate_list))):
            # log.forcedLog(f"coordinate should be larger than 0 : erorr {coorinate_list}")
        if n >= 1:
            self._row = n
            self.cell_active = self.ws_active.cell(row=self._row, column=self._col)
        else:
            log_wrapping("excel_log", f"row number should be larger than 0 : erorr {n}")    


    @property
    def column(self):
        return self._col
    

    @column.setter
    def column(self, n):

        # if len(list(filter(lambda x: x <= 0, coorinate_list))):
            # log.forcedLog(f"coordinate should be larger than 0 : erorr {coorinate_list}")
        if n >= 1:
            self._col = n
            self.cell_active = self.ws_active.cell(row=self._row, column=self._col)
        else:
            log_wrapping("excel_log", f"column number should be larger than 0 : erorr {n}") 

    @property
    def position(self):

        print(f"[{self.filename}] row={self._row}, column={self._col}")
        return [self._row, self._col]
    

    @position.setter
    def position(self, *args):

        self.row = args[0][0]
        self.column = args[0][1]


    @property
    def insert_row(self):

        # insert empty row at self._row
        log_wrapping("excel_log", f"insert a row at {self._row}")
        self.ws_active.insert_rows(self._row)
        self.save
    

    @property
    def insert_column(self):

        # insert empty column at self._col
        log_wrapping("excel_log", f"insert a column at {self._col}")
        self.ws_active.insert_cols(self._col)
        self.save
    

    def merge_cell(self, start_row, start_column, end_row, end_column):

        self.ws_active.merge_cells(start_row, start_column, end_row, end_column)

    
    @property
    def image(self):

        ret = self._image
        return ret


    @image.setter
    def image(self, img):
        
        # img format : [object, width, height]
        # 100 = 2.65cm

        self._image = Image(img[0])
        self._image.width  = img[1]
        self._image.height = img[2]
        coordinate = self.cell_active.coordinate
        self.ws_active.add_image(self._image, coordinate)
        self.save
    

    @property
    def autosize_column(self):

        for i, col in enumerate(self.ws_active.columns):
            length = max(len(str(cell.value)) for cell in col)
            self.ws_active.column_dimensions[col[0].column_letter].width = length + 1
        self.save
    

    @property
    def cell_font(self):

        self.update_font
        header = ["font item", "value"]
        ret_map = [header] + [
            ["row", self._row],
            ["column", self._col],
            ["name", self.set_font_type],
            ["size", self.set_font_size],
            ["bold", self.set_font_bold],
            ["italic", self.set_font_italic],
            ["vertical position", self.set_font_vert],
            ["underline", self.set_font_underline],
            ["strike", self.set_font_strike],
            ["color", self.set_font_color]
            ]
        
        print(tb(ret_map, headers="firstrow"))
        self.save
    

    @cell_font.setter
    def cell_font(self, value=style.context):

        self.cell_active.font = value
        self.update_font

    
    @property
    def update_font(self):

        self.cell_active.font = Font(
            name      = self.set_font_type,
            size      = self.set_font_size,
            bold      = self.set_font_bold,
            italic    = self.set_font_italic,
            vertAlign = self.set_font_vert,
            underline = self.set_font_underline,
            strike    = self.set_font_strike,
            color     = self.set_font_color
        )


    @property
    def cell_color(self):
        pass

    
    @cell_color.setter
    def cell_color(self, rgb=color_hex.white):
        
        self.set_cell_color = rgb
        self.cell_active.fill = PatternFill(
            fill_type = "solid",
            fgColor   = Color(self.set_cell_color)
        )
    

    @property
    def border(self):
        pass

    
    @border.setter
    def border(self, color=color_hex.black):
        
        self.cell_active.border = Border(
            left   = Side(border_style=self._border, color=color),
            right  = Side(border_style=self._border, color=color),
            top    = Side(border_style=self._border, color=color),
            bottom = Side(border_style=self._border, color=color)
        )
    

    @property
    def border_none(self):

        self.cell_active.border = Border(
            left   = Side(border_style=None),
            right  = Side(border_style=None),
            top    = Side(border_style=None),
            bottom = Side(border_style=None)
        )
    

    @property
    def vertical_align(self):
        pass

    
    @vertical_align.setter
    def vertical_align(self, position):
        
        self._align_vertical = position
        self.cell_active.alignment = Alignment(
            vertical=self._align_vertical, horizontal=self._align_horizontal)
    

    @property
    def horizontal_align(self):
        pass

    
    @horizontal_align.setter
    def horizontal_align(self, position):
        
        self._align_horizontal = position
        self.cell_active.alignment = Alignment(
            vertical=self._align_vertical, horizontal=self._align_horizontal)
    

    @property
    def cell_height(self):
        pass


    @cell_height.setter
    def cell_height(self, n):

        # 100 = 133px
        # default 15.03 = 20px

        self.ws_active.row_dimensions[self._row].height = n / 1.33
    

    @property
    def cell_width(self):
        pass

    
    @cell_width.setter
    def cell_width(self, var):

        # if n is 10 = 91px
        # default cell height is 17px 
        # require a alphabet in the column
        # if n is 1.0989 = 10px

        self.ws_active.column_dimensions[var[0]].width = var[1] * 0.14285
    
    
    @property
    def set_header(self):

        # self.update_font is not required
        self.cell_active.font = style.header
        self.cell_color = color_hex.dark_silver


    @property
    def set_context(self):

        # self.update_font is not required
        self.cell_active.font = style.context
        self.cell_color = color_hex.white
    

    @property
    def insert_header(self):
        pass


    @insert_header.setter
    def insert_header(self, *args):

        row = args[0][0]
        col = args[0][1]
        variable = args[0][2]
        if isinstance(variable, list):
            for cell_column in range(col, (len(variable)+col)):
                self.position = row, cell_column
                self.set_header
                self.cell = variable[cell_column-col]
            log_wrapping("excel_log", f"insert header to row {row}")
        else:
            log_wrapping("excel_log", f"instance error, variable shall be list type")
    

    @property
    def insert_list(self):
        pass


    @insert_list.setter
    def insert_list(self, *args):
        
        row = args[0][0]
        col = args[0][1]
        variable = args[0][2]
        if isinstance(variable, list):
            for cell_column in range(col, (len(variable)+col)):
                self.position = row, cell_column
                self.border = color_hex.dark_silver
                self.set_context
                self.cell = variable[cell_column-col]
        else:
            log_wrapping("excel_log", f"instance error, variable shall be list type")