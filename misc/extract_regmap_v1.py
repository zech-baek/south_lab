# - purpose
#    generate the register map and map for status printing


from openpyxl import load_workbook
from interface.cui_colors import color
from interface.cui_logger import logger as log
import os
import re
import yaml

log.initLogger(log.info)



class extract_regmap():
    
    '''
    usage example:
        from misc.extract_regmap import extract_regmap
        hl7133 = extract_regmap(file="./misc/hl7133_ae_regmap.xlsx")
        hl7133.sorting_regmap()
    '''

    def __init__(self, file) -> None:

        self.filename   = file.replace(".xlsx", "")
        self.excel      = load_workbook(file, data_only=True)
        self.sheet_list = self.excel.sheetnames
        self.working    = self.excel[self.sheet_list[0]]

        # self.excel.calculate_all() # force to recalculation of formulas in whole cells

        self.column = {
            1  : "address",
            2  : "register",
            3  : "group",
            4  : "rw",
            5  : "por",
            6  : "bit7",
            7  : "bit6",
            8  : "bit5",
            9  : "bit4",
            10 : "bit3",
            11 : "bit2",
            12 : "bit1",
            13 : "bit0"
        }

        self.regmap_structure = ["name", "register", "address", "group", "rw", "msb", "lsb", 'link']

        # empty cell : NoneType
        # if there merged cell, only the first cell has a value, otherwise returns None

        self.max_row = 0
        self.max_column = 0


    def _check_maxCell(self):

            for row in range(1, 1000):
                res_row = self.working.cell(row=row, column=1).value

                if res_row == None:
                    log.infoLog(f"last row of the cell : {row-1}")
                    self.max_row = row-1
                    break
            for col in range(1, 20):
                res_col = self.working.cell(row=1, column=col).value
                if res_col == None:
                    log.infoLog(f"last column of the cell : {col-1}")
                    self.max_column = col-1
                    break
            

    def sorting_regmap(self):

        '''
        example format of register.yaml
            ADC_IBAT:
                address: 36
                group: user
                lsb: 0
                msb: 3
                name: ADC_IBAT
                register: ADC_IBAT_1
                rw: R
        
        example format of register_status.yaml
            163:
                - TESTREG3          # register
                - scan_mode         # bit 7
                - en_atbpin0        # bit 6
                - en_dtbpin         # bit 5
                - tst_adc_afa_en    # bit 4
                - tst_adc_chop_sign # bit 3
                - qrbilim           # bit 2
                - tst_bgpol         # bit 1
                - en_ckext          # bit 0
        '''

        self._check_maxCell()

        self.regmap = dict()
        self.regmap_status = dict()

        for row in range(2, self.max_row+1):

            addr_raw = self.working.cell(row=row, column=1).value
            if isinstance(addr_raw, int):
                address = addr_raw
            elif isinstance(addr_raw, str):
                address = int(addr_raw, 16)
            register = self.working.cell(row=row, column=2).value
            group    = self.working.cell(row=row, column=3).value
            rw       = self.working.cell(row=row, column=4).value

            bits_list = list()
            bits_list.append(register)
            temp_name = None
            temp_link = None
            
            log.infoLog(f"{color.blue}{address} {register} {group} {rw}{color.end}")
            
            for col in range(6, self.max_column+1):
                
                # read the cell from bit 7 to 0
                read_cell = self.working.cell(row=row, column=col).value
                
                if read_cell != None:
                    
                    trimmed_name = self._trim_string(read_cell)
                    link_check = self._get_number(read_cell)
                    temp_name = trimmed_name
                    temp_link = link_check
                    
                    if temp_link == None:
                        # single register
                        if trimmed_name in self.regmap.keys():
                            log.infoLog(f"{trimmed_name} already exists")
                            pass
                        else:
                            self.regmap[temp_name] = dict()
                        link = {
                            'low_reg' : address,
                            'low_lsb' : 13-col,
                            'low_msb' : 13-col,
                            'high_reg' : None,
                            'high_lsb' : None,
                            'high_msb' : None
                        }
                        self.regmap[temp_name]['multi'] = 0
                        self.regmap[temp_name]['link'] = link
                        log.infoLog(f"case 1 : reg-{temp_name} link-{temp_link}, multi-{self.regmap[temp_name]['multi']}, low reg-{link['low_reg']} lsb-{link['low_lsb']} msb-{link['low_msb']}, high reg-{link['high_reg']} lsb-{link['high_lsb']} msb-{link['high_msb']}")
                    else:
                        if trimmed_name in self.regmap.keys():
                            # multi-bit & update date if exist

                            self.regmap[temp_name]['multi'] = 1

                            if temp_link[1] != 0:
                                self.regmap[temp_name]['link']['high_reg'] = address
                                self.regmap[temp_name]['link']['high_lsb'] = 13-col
                                self.regmap[temp_name]['link']['high_msb'] = 13-col
                                log.infoLog(f"case 2-1 : reg-{temp_name} link-{temp_link}, multi-{self.regmap[temp_name]['multi']}, low reg-{self.regmap[temp_name]['link']['low_reg']} lsb-{self.regmap[temp_name]['link']['low_lsb']} msb-{self.regmap[temp_name]['link']['low_msb']}, high reg-{self.regmap[temp_name]['link']['high_reg']} lsb-{self.regmap[temp_name]['link']['high_lsb']} msb-{self.regmap[temp_name]['link']['high_msb']}")
                            else:
                                self.regmap[temp_name]['link']['low_reg'] = address
                                self.regmap[temp_name]['link']['low_lsb'] = 13-col
                                self.regmap[temp_name]['link']['low_msb'] = 13-col
                                log.infoLog(f"case 2-2 : reg-{temp_name} link-{temp_link}, multi-{self.regmap[temp_name]['multi']}, low reg-{self.regmap[temp_name]['link']['low_reg']} lsb-{self.regmap[temp_name]['link']['low_lsb']} msb-{self.regmap[temp_name]['link']['low_msb']}, high reg-{self.regmap[temp_name]['link']['high_reg']} lsb-{self.regmap[temp_name]['link']['high_lsb']} msb-{self.regmap[temp_name]['link']['high_msb']}")
                        else:
                            # multi-bit & generate data if not exist
                            self.regmap[temp_name] = dict()
                            self.regmap[temp_name]['multi'] = 0
                            if temp_link[1] != 0:
                                link = {
                                    'low_reg' : None,
                                    'low_lsb' : None,
                                    'low_msb' : None,
                                    'high_reg' : address,
                                    'high_lsb' : 13-col,
                                    'high_msb' : 13-col
                                }
                                log.infoLog(f"case 3 : reg-{temp_name} link-{temp_link}, multi-{self.regmap[temp_name]['multi']}, low reg-{link['low_reg']} lsb-{link['low_lsb']} msb-{link['low_msb']}, high reg-{link['high_reg']} lsb-{link['high_lsb']} msb-{link['high_msb']}")
                            else:
                                link = {
                                    'low_reg' : address,
                                    'low_lsb' : 13-col,
                                    'low_msb' : 13-col,
                                    'high_reg' : None,
                                    'high_lsb' : None,
                                    'high_msb' : None
                                }
                                log.infoLog(f"case 4 : reg-{temp_name} link-{temp_link}, multi-{self.regmap[temp_name]['multi']}, low reg-{link['low_reg']} lsb-{link['low_lsb']} msb-{link['low_msb']}, high reg-{link['high_reg']} lsb-{link['high_lsb']} msb-{link['high_msb']}")
                            self.regmap[temp_name]['link'] = link
                else:
                    # read_cell is None
                    if temp_link != None:
                        # multi-bit
                        if temp_link[1] != 0:
                            self.regmap[temp_name]['link']['high_lsb'] = 13-col
                            log.infoLog(f"case 5 : reg-{temp_name} link-{temp_link}, multi-{self.regmap[temp_name]['multi']}, low reg-{self.regmap[temp_name]['link']['low_reg']} lsb-{self.regmap[temp_name]['link']['low_lsb']} msb-{self.regmap[temp_name]['link']['low_msb']}, high reg-{self.regmap[temp_name]['link']['high_reg']} lsb-{self.regmap[temp_name]['link']['high_lsb']} msb-{self.regmap[temp_name]['link']['high_msb']}")
                        else:
                            self.regmap[temp_name]['link']['low_lsb'] = 13-col
                            log.infoLog(f"case 6 : reg-{temp_name} link-{temp_link}, multi-{self.regmap[temp_name]['multi']}, low reg-{self.regmap[temp_name]['link']['low_reg']} lsb-{self.regmap[temp_name]['link']['low_lsb']} msb-{self.regmap[temp_name]['link']['low_msb']}, high reg-{self.regmap[temp_name]['link']['high_reg']} lsb-{self.regmap[temp_name]['link']['high_lsb']} msb-{self.regmap[temp_name]['link']['high_msb']}")
                    if temp_name == "RSVD":
                        self.regmap[temp_name]['link']['low_lsb'] = 13-col
                        log.infoLog(f"case 7 : reg-{temp_name} link-{temp_link}, multi-{self.regmap[temp_name]['multi']}, low reg-{self.regmap[temp_name]['link']['low_reg']} lsb-{self.regmap[temp_name]['link']['low_lsb']} msb-{self.regmap[temp_name]['link']['low_msb']}, high reg-{self.regmap[temp_name]['link']['high_reg']} lsb-{self.regmap[temp_name]['link']['high_lsb']} msb-{self.regmap[temp_name]['link']['high_msb']}")

                bits_list.append(temp_name)
                self.regmap[temp_name]["name"    ] = temp_name
                self.regmap[temp_name]["address" ] = address  
                self.regmap[temp_name]["register"] = register 
                self.regmap[temp_name]["group"   ] = group    
                self.regmap[temp_name]["rw"      ] = rw
                
                # log.infoLog(f"{temp_name} {address} {register} {group} {rw} {self.regmap[temp_name]['link']['low_msb']} {self.regmap[temp_name]['link']['low_lsb']} {self.regmap[temp_name]['link']['high_msb']} {self.regmap[temp_name]['link']['high_lsb']}")
            # log.infoLog(f"{color.blue}{address} {register} {group} {rw}{color.end}")
            self.regmap_status[address] = bits_list

        file1 = open(f"{self.filename}_reg.yaml", "w")
        yaml.dump(self.regmap, file1, default_flow_style=False)
        log.infoLog(f"dump the register map to {self.filename}_reg.yaml")

        file2 = open(f"{self.filename}_status.yaml", "w")
        yaml.dump(self.regmap_status, file2, default_flow_style=False)
        log.infoLog(f"dump the status map to {self.filename}_status.yaml")

        file1.close()
        file2.close()
        
        
    def _get_number(self, string):
        
        pattern1 = r"<(\d+:\d+)>"
        pattern2 = r"\[(\d+:\d+)\]"
        pattern3 = r"\[(\d+)\]"
        pattern4 = r"<(\d+)>"
        match1 = re.search(pattern1, string)
        match2 = re.search(pattern2, string)
        match3 = re.search(pattern3, string)
        match4 = re.search(pattern4, string)
        
        if match1:
            extract_num = match1.group(1)
        elif match2:
            extract_num = match2.group(1)
        elif match3:
            extract_num = match3.group(1)
        elif match4:
            extract_num = match4.group(1)
        else:
            return None
        
        if ":" in extract_num:
            ret = [int(num) for num in extract_num.split(":")]
        else:
            ret = [int(extract_num), int(extract_num)]
        
        return ret
    

    def _trim_string(self, string):

        '''
        remove white space and empty spaces
            1. tab, cr and spaces
            2. numbers on the head
            3. special characters
            4. msb and lsb information at the tail
        '''
        
        pattern_remove_1 = re.sub(r"\[.*?\]", "", string) # Remove everything between '[' and ']'
        pattern_remove_2 = re.sub(r"<.*", "", pattern_remove_1) # Remove everything after '<'
        pattern_remove_3 = re.sub(r"\s+", "_", pattern_remove_2) # Replace multiple spaces, tabs, and carriage returns with a single underscore
        pattern_remove_4 = re.sub(r"^\d+", "", pattern_remove_3) # Remove leading digits
        pattern_remove_5 = re.sub(r"[^\w\s-]", "", pattern_remove_4) # Remove all special characters
        
        if pattern_remove_1 != pattern_remove_5:
            log.forcedLog(f"modify the string from {pattern_remove_1} to {color.bgyel}{pattern_remove_5}{color.end}")
        
        return pattern_remove_5