# ! /usr/bin/env python
# coding=utf-8

import os
import sys
import pathlib

try:
    # try to use __file__
    misc_dir = os.path.dirname(os.path.abspath(__file__))
except NameError:
    # fallback to sys.argv[0] or current working directory
    if len(sys.argv) > 0:
        misc_dir = os.path.dirname(os.path.abspath(sys.argv[0]))
    else:
        misc_dir = os.getcwd()

log_dir = pathlib.Path(misc_dir).parent/"log"

if not log_dir.exists():
    log_dir.mkdir(parents=True, exist_ok=True)


from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from interface.cui_colors import color
from interface.cui_logger import logger as log
import re, yaml

log.initLogger(log.info)


class extract_regmap():
    
    '''
    e.g. :
        file = "./misc/regmap/sc8583_1p1_regmap.xlsx"
        
        from misc.extract_regmap_v2 import extract_regmap
        target_object = extract_regmap(filename=file, project="sc8583", revision="1p1")
        target_object.extract()
    '''

    def __init__(self, filename, project, revision) -> None:

        self.path       = os.getcwd()
        self.project    = project
        self.revision   = revision
        self.excel      = load_workbook(filename)
        self.sheet_list = self.excel.sheetnames
        self.sheet      = self.excel[self.sheet_list[0]]

        
    def extract(self):
        
        self.regmap = dict()
        self.stsmap = dict()
        self.tablemap = dict()

        # get the range for merged cells
        # e.g.
        #   [<MergedCellRange H53:I53>,
        #   <MergedCellRange F78:M78>,
        #   <MergedCellRange F87:M87>,
        #   <MergedCellRange F26:K26>,
        #   <MergedCellRange F128:M128>]

        self.merged_ranges = list(self.sheet.merged_cells.ranges)
        
        for row in self.sheet.iter_rows(min_row=2):
            
            # variable row in for loop returns tuple which include the parameter from A_min_row to valid cell
            # e.g.
            # (<Cell 'Sheet1'.A2>, <Cell 'Sheet1'.B2>, <Cell 'Sheet1'.C2>, <Cell 'Sheet1'.D2>, <Cell 'Sheet1'.E2>, <Cell 'Sheet1'.F2>, <MergedCell 'Sheet1'.G2>, <MergedCell 'Sheet1'.H2>, <MergedCell 'Sheet1'.I2>, <MergedCell 'Sheet1'.J2>, <MergedCell 'Sheet1'.K2>, <MergedCell 'Sheet1'.L2>, <MergedCell 'Sheet1'.M2>)
            
            # log.forcedLog(f"row_0={row[0].value}, reg={row[1].value}, perm={row[2].value}, rw={row[3].value}, por={row[4].value}")
            
            if row[0].value is not None:
                address    = int(row[0].value, 0) # auto detect
                register   = self._trim_string(string=row[1].value)
                permission = row[2].value
                rw         = row[3].value
                por        = int(row[4].value, 16)
                
                # form for list_register : [register, name_bit7, name_bit6, ... name_bit0]
                list_register  = list()
                list_register.append(register)
                
                # enumerate bit7 ~ bit0

                for col_idx in range(5, 13):
                    
                    cell = row[col_idx]

                    # return the cell coordinate only : e.g. F5
                    cell_address = cell.coordinate

                    # return merged cell range : e.g. F2:M2 or None
                    merged_range = next((r for r in self.merged_ranges if cell_address in r), None)

                    # splited or multi position register
                    if merged_range:

                        # return the tuple : e.g. (6, 2, 13, 2)
                        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
                        
                        name_raw   = self.sheet.cell(row=min_row, column=min_col).value
                        if "RSVD" not in name_raw and "RESERVED" not in name_raw:
                            
                            # split the name to list : e.g. id[13:6] -> id$13$6 -> ["id", "13", "6"]
                            text_split = name_raw.replace("[", "$").replace("]", "").replace(":", "$")
                            split_list = text_split.split("$")
                            name       = self._trim_string(string=split_list[0])
                            bit_end    = int(split_list[1])
                            bit_start  = int(split_list[2])
                        
                            if name not in self.regmap:

                                self.regmap[name] = {
                                    "name"       : name,
                                    "rw"         : rw,
                                    "permission" : permission,
                                    "reset"      : por,
                                    "singular"   : True
                                }
                            
                            entry = self.regmap[name]

                            # if not exist in the self.regmap[name], return set())
                            # if exist, return : e.g. (3, 13, 6) for address 3 and [13:6]
                            existing_combinations = {
                                (entry[f"address{i}"], entry[f"bith{i}"], entry[f"bitl{i}"])
                                for i in range(1, len([k for k in entry.keys() if k.startswith("address")]) + 1)
                                }
                            new_combination = (address, bit_end, bit_start)
                            
                            if new_combination not in existing_combinations:

                                # obtaion the length for splited name block 
                                index = len([k for k in entry.keys() if k.startswith("address")]) +1
                                entry[f"address{index}"]  = address
                                entry[f"register{index}"] = register
                                entry[f"lsb{index}"]      = 13-max_col # actual position on the byte
                                entry[f"msb{index}"]      = 13-min_col # actual position on the byte
                                entry[f"bith{index}"]     = bit_end    # logical position
                                entry[f"bitl{index}"]     = bit_start  # logical position
                            
                            if  index > 1:
                                font_color = color.blue
                            else:
                                font_color = color.cyan
                            log.forcedLog(f"{font_color}addr{index} {address:#04x} {name}[{bit_end}:{bit_start}]{color.end} : msb {13-min_col}, lsb {13-max_col}")
                            
                            if (bit_end - bit_start) != 0:
                                entry[f"singular"] = False
                        
                        list_register.append(name)
                    
                    # singular register
                    else:

                        if cell.value:
                            bit_position = 7 - (col_idx-5)

                            if "[" in cell.value:
                                raise Exception(f"format error : {cell.value}")
                            
                            name = self._trim_string(string=cell.value)

                            if "RSVD" not in name and "RESERVED" not in name:

                                if name not in self.regmap:

                                    self.regmap[name] = {
                                        "name"       : cell.value,
                                        "rw"         : rw,
                                        "permission" : permission,
                                        "reset"      : por,
                                        "singular"   : True
                                    }

                                    entry = self.regmap[name]
                                    index = len([k for k in entry.keys() if k.startswith("address")]) + 1
                                    entry[f"address{index}"]  = address
                                    entry[f"register{index}"] = register
                                    entry[f"msb{index}"]      = bit_position # actual position on the byte
                                    entry[f"lsb{index}"]      = bit_position # actual position on the byte
                                    entry[f"bith{index}"]     = 0 # logical position
                                    entry[f"bitl{index}"]     = 0 # logical position

                                log.forcedLog(f"{color.yellow}addr{index} {address:#04x} {name}{color.end} : msb {bit_position}, lsb {bit_position}")
                            
                            list_register.append(name)

            self.stsmap[address] = list_register

        file1 = open(log_dir/f"{self.project}_{self.revision}_reg.yaml", "w")
        yaml.dump(self.regmap, file1, default_flow_style=False)
        log.infoLog(f"dump the register map to yaml")

        file2 = open(log_dir/f"{self.project}_{self.revision}_status.yaml", "w")
        yaml.dump(self.stsmap, file2, default_flow_style=False)
        log.infoLog(f"dump the status map to yaml")

        file1.close()
        file2.close()
    

    def _trim_string(self, string):

        '''
        remove white space and empty spaces
            1. tab, cr and spaces
            2. numbers on the head
            3. special characters
            4. msb and lsb information at the tail
        '''
        if string.startswith(" "):
            raise Exception(f"format error : {string}")
        
        pattern_remove_1 = re.sub(r"\[.*?\]", "", string) # Remove everything between '[' and ']'
        pattern_remove_2 = re.sub(r"<.*", "", pattern_remove_1) # Remove everything after '<'
        pattern_remove_3 = re.sub(r"\s+", "", pattern_remove_2) # Replace multiple spaces, tabs, and carriage returns
        pattern_remove_4 = re.sub(r"^\d+", "", pattern_remove_3) # Remove leading digits
        pattern_remove_5 = re.sub(r"[^\w\s-]", "", pattern_remove_4) # Remove all special characters
        
        if pattern_remove_1 != pattern_remove_5:
            log.forcedLog(f"modify the string from {color.red}{pattern_remove_1}{color.end} to {color.bgyel}{color.red}{pattern_remove_5}{color.end}")
        
        return pattern_remove_5