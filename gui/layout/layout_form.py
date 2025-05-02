'''
1. use the template_tab2.xlsx file
2. change the register map sheet using "project.get_device_info"
    e.g.
        from project.get_device_info import get_excel_regmap
        get_excel_regmap(device="xx", revision="aa")
'''


import warnings
warnings.simplefilter("ignore", UserWarning)

import dearpygui.dearpygui as dpg
from openpyxl import load_workbook


class tab_form:
    
    def __init__(self, file, reg_map):
        
        self.wb = load_workbook(file)
        self.sheet_list = ['Guide', 'GroupForm', 'ValueForm', 'Register map']
        self.sheet = self.wb[self.sheet_list[1]]
        
        self.reg_map = reg_map
        self.layout_table = None
        self.group_a = None
        self.group_b = None
    
    
    def close(self):
        
        self.wb = None
        self.sheet_list = None
        self.sheet = None
        self.reg_map = None
        self.layout_table = None
        self.group_a = None
        self.group_b = None
    
    
    def __del__(self):
        pass
        
        
    def get_layout_table(self):
        
        self.extract_form()
        self.get_length_rw()
        self.divide_group()
        
        self.close_excel()
        
        return self.group_a, self.group_b
    
    
    def close_excel(self):
        self.wb.close()
        
        
    def extract_form(self):
        
        '''
        cell data according to (row, column)
        (x, 2) Main-Group : a or b
        (x, 3) Sub-Group : title name of the table
        (x, 4) Register : register name
        (x, 5) selection : selection option, radio button, textbox and dropdown menu are support in the gui
        (x, 6~) value : text beside the selection
            style 1. "int"
            style 2. "int, comment"
            style 3. None : simply generate the corresponding comment based on the bits
        '''

        table_property = dict()
        
        x = 3
        y = 2

        while True:
            
            # check the 2nd column data
            if self.sheet.cell(row=x, column=y).value is None:
                break
            
            else:
                list_style = list()
                
                while True:
                    
                    ret_cell = self.sheet.cell(row=x, column=y).value
                    
                    if   y == 2: main_group = ret_cell
                    elif y == 3: sub_group  = ret_cell
                    elif y == 4: register   = ret_cell
                    elif y == 5: selection  = ret_cell
                    elif y >= 6:
                        if ret_cell is None:
                            break
                        else:
                            list_style.append(ret_cell)
                        
                    y += 1
                
                table_property[register] = {
                    "main_group": main_group,
                    "sub_group" : sub_group,
                    "register"  : register,
                    "selection" : selection,
                    "style"     : list_style
                }
                    
                y = 2        
                x += 1
    
        self.layout_table = table_property

    
    def get_length_rw(self):
        
        for key in self.layout_table.keys():
        
            total_length = 0
            
            # find the number that starts with "lsb" to calculate the total bit length
            parsing_length = len([item for item in self.reg_map[key].keys() if item.startswith("lsb")])
            for index in range(1, parsing_length+1):
                total_length += self.reg_map[key][f"msb{index}"] - self.reg_map[key][f"lsb{index}"] + 1

            self.layout_table[key].update({"length" : total_length})
            self.layout_table[key].update({"rw" : self.reg_map[key]["rw"]})
    
    
    # divide the all items into group a and b
    def divide_group(self):
        
        self.group_a = {k: v for k, v in self.layout_table.items() if v.get("main_group") == "a"}
        self.group_b = {k: v for k, v in self.layout_table.items() if v.get("main_group") == "b"}
    
    
    # get the table title only
    def get_sub_group(self):
        
        sub_group_a = list(set(self.group_a[key]["sub_group"] for key in self.group_a))
        sub_group_b = list(set(self.group_b[key]["sub_group"] for key in self.group_b))
        
        return sub_group_a, sub_group_b
    
    
    # get the dictionary for key=table title, value=register name only
    def get_table_item(self):
        
        sub_group_a, sub_group_b = self.get_sub_group()
        
        group_item_a = dict()
        group_item_b = dict()

        for table_title in sub_group_a:
            temp_a = [k for k, v in self.group_a.items() if v.get("sub_group") == table_title]
            group_item_a.update({table_title: temp_a})
        for table_title in sub_group_b:
            temp_b = [k for k, v in self.group_b.items() if v.get("sub_group") == table_title]
            group_item_b.update({table_title: temp_b})
        
        return group_item_a, group_item_b